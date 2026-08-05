import datetime as dt
import json
import os
import re
import sys
import unicodedata
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


def normalized_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().casefold()


def json_value(value):
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def resolve_column(value, headers, max_column):
    wanted = str(value or "").strip()
    if not wanted:
        return None
    if re.fullmatch(r"[A-Za-z]{1,3}", wanted):
        index = column_index_from_string(wanted.upper())
        return index if index <= max_column else None
    needle = normalized_text(wanted)
    for index, header in headers.items():
        if normalized_text(header) == needle:
            return index
    return None


def month_matches(value, month, year=None):
    if month is None:
        return True
    month = int(month)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.month == month and (year is None or value.year == int(year))
    text = normalized_text(value)
    if not text:
        return False
    year_match = re.search(r"\b(20\d{2})\b", text)
    if year is not None and year_match and int(year_match.group(1)) != int(year):
        return False
    patterns = [
        rf"\bthang\s*0?{month}\b",
        rf"\bt\s*0?{month}\b",
        rf"\b0?{month}\s*[/.-]\s*(?:20\d{{2}}|\d{{2}})\b",
        rf"\b\d{{1,2}}\s*[/.-]\s*0?{month}(?:\s*[/.-]\s*(?:20\d{{2}}|\d{{2}}))?\b",
    ]
    return any(re.search(pattern, text) for pattern in patterns) or text in {str(month), f"{month:02d}"}


def normalized_url(value):
    text = str(value or "").strip()
    if not re.match(r"^https?://", text, re.I):
        return text.casefold()
    try:
        parts = urlsplit(text)
        query = [
            (key, item)
            for key, item in parse_qsl(parts.query, keep_blank_values=True)
            if not key.casefold().startswith("utm_")
        ]
        path = parts.path.rstrip("/") or "/"
        return urlunsplit((parts.scheme.casefold(), parts.netloc.casefold(), path, urlencode(query), ""))
    except ValueError:
        return text.casefold()


def main():
    request = json.load(sys.stdin)
    path = os.path.abspath(str(request.get("path") or ""))
    options = request.get("options") or {}
    if not os.path.isfile(path) or not path.casefold().endswith((".xlsx", ".xlsm")):
        raise ValueError("Workbook path is missing or unsupported.")

    workbook = load_workbook(path, read_only=False, data_only=False)
    requested_sheet = str(options.get("sheet") or "").strip()
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {requested_sheet}")
        sheet = workbook[requested_sheet]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    header_row = max(1, int(options.get("header_row") or 1))
    headers = {
        index: (sheet.cell(header_row, index).value or get_column_letter(index))
        for index in range(1, sheet.max_column + 1)
    }
    filter_column = resolve_column(options.get("filter_column"), headers, sheet.max_column)
    link_column = resolve_column(options.get("link_column"), headers, sheet.max_column)
    if options.get("filter_column") and filter_column is None:
        raise ValueError(f"Filter column not found: {options.get('filter_column')}")
    if options.get("link_column") and link_column is None:
        raise ValueError(f"Link column not found: {options.get('link_column')}")

    required = []
    for value in options.get("required_columns") or []:
        index = resolve_column(value, headers, sheet.max_column)
        if index is None:
            raise ValueError(f"Required column not found: {value}")
        required.append(index)
    range_start = resolve_column(options.get("required_range_start"), headers, sheet.max_column)
    range_end = resolve_column(options.get("required_range_end"), headers, sheet.max_column)
    if range_start:
        range_end = range_end or sheet.max_column
        required.extend(range(range_start, range_end + 1))
    required = sorted(set(required))

    month = options.get("filter_month")
    year = options.get("filter_year")
    max_rows = min(5000, max(1, int(options.get("max_rows") or 1000)))
    rows = []
    links = {}
    total_matched = 0
    for row_number in range(header_row + 1, sheet.max_row + 1):
        if filter_column and not month_matches(sheet.cell(row_number, filter_column).value, month, year):
            continue
        total_matched += 1
        if len(rows) >= max_rows:
            continue
        values = {}
        for index, header in headers.items():
            value = sheet.cell(row_number, index).value
            if value not in (None, ""):
                values[str(header)] = json_value(value)
        missing = [str(headers[index]) for index in required if sheet.cell(row_number, index).value in (None, "")]
        link = ""
        if link_column:
            cell = sheet.cell(row_number, link_column)
            hyperlink = getattr(cell, "hyperlink", None)
            link = str(hyperlink.target if hyperlink else cell.value or "").strip()
        key = normalized_url(link)
        if key:
            links.setdefault(key, {"url": link, "rowNumbers": []})["rowNumbers"].append(row_number)
        rows.append({
            "rowNumber": row_number,
            "link": link,
            "missingRequired": missing,
            "values": values,
        })

    duplicate_links = [item for item in links.values() if len(item["rowNumbers"]) > 1]
    result = {
        "object": "browser.workbook.audit",
        "data": {
            "filename": os.path.basename(path),
            "sheet": sheet.title,
            "sheetNames": workbook.sheetnames,
            "headerRow": header_row,
            "headers": [{"column": get_column_letter(index), "name": str(value)} for index, value in headers.items()],
            "filter": {
                "column": get_column_letter(filter_column) if filter_column else None,
                "month": month,
                "year": year,
            },
            "totalMatchedRows": total_matched,
            "returnedRows": len(rows),
            "rowsTruncated": total_matched > len(rows),
            "duplicateLinks": duplicate_links,
            "uniqueLinks": list(links.values()),
            "rows": rows,
        },
    }
    json.dump(result, sys.stdout, ensure_ascii=False, separators=(",", ":"))


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        json.dump({"error": {"type": type(error).__name__, "message": str(error)}}, sys.stdout, ensure_ascii=False)
        sys.exit(1)
